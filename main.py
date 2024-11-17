import json
import http.client
import geopandas as gpd
from shapely.geometry import Point, LineString
import os
import time
import socket
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# Toggle to enable/disable testing a specific job
TEST_ONLY_SPECIFIC_JOB = False

# ID of the specific job to test
TEST_JOB_ID = "-O-nlOLQbPIYhHwJCPDN"
API_KEY = 'rt2JR8Rds03Ry03hQTpD9j0N01gWEULJnuY3l1_GeXA8uqUVLtXsKHUQuW5ra0lt-FklrA40qq6_J04yY0nPjlfKG1uPerclUX2gf6axkIioJadYxzOG3cPZJLRcZ2_vHPdipZWvQdICAL2zRnqnOUCGjfq4Q8aMdmA7H6z7xK7W9MEKnIiEALokmtChLtr-s6hDFko17M7xihPpNlfGN7N8D___wn55epkLMtS2eFF3JPlj_SjpFIGXYK15PJFta-BmPqCFvEwXlZEYfEf8uYOpAvCEdBn3NOMoB-P28owOJ7ZeBQf5VMFi3J5_RV2fE_XDR2LTD469Qq0y3946LQ'

# Function to get list of jobs from KatapultPro API
def getJobList():
    URL_PATH = '/api/v2/jobs'
    headers = {}
    all_jobs = []

    for attempt in range(5):  # Consider using exponential backoff to avoid overwhelming the server
        conn = None
        try:
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=10)  # Timeout value could be made configurable
            conn.request("GET", f"{URL_PATH}?api_key={API_KEY}", headers=headers)
            res = conn.getresponse()
            data = res.read().decode("utf-8")
            jobs_dict = json.loads(data)

            if not isinstance(jobs_dict, dict):
                raise TypeError(f"Expected a dictionary but received {type(jobs_dict)}: {jobs_dict}")

            # Retrieve all jobs without filtering for status
            all_jobs = [
                {'id': job_id, 'name': job_details.get('name'), 'status': job_details.get('status')}
                for job_id, job_details in jobs_dict.items()
            ]
            break

        except (socket.error, OSError) as e:
            print(f"Socket error: {e}. Retrying...")
            time.sleep(5)
        except Exception as e:
            print(f"Failed to retrieve job list: {e}")
            break
        finally:
            if conn:
                conn.close()

    return all_jobs

# Function to get job data from KatapultPro API
def getJobData(job_id):
    URL_PATH = f'/api/v2/jobs/{job_id}'
    headers = {}

    for attempt in range(5):  # Consider using exponential backoff to avoid overwhelming the server
        conn = None
        try:
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=10)  # Timeout value could be made configurable
            conn.request("GET", f"{URL_PATH}?api_key={API_KEY}", headers=headers)
            res = conn.getresponse()
            data = res.read().decode("utf-8")
            job_data = json.loads(data)

            if "error" in job_data and job_data["error"] == "RATE LIMIT EXCEEDED":
                print("Rate limit exceeded. Retrying after delay...")
                time.sleep(5)
                continue
                # Save job data to a file if testing a specific job
            if TEST_ONLY_SPECIFIC_JOB:
                workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
                file_path = os.path.join(workspace_path, f"test_job_{job_id.replace('/', '_')}.json")
                with open(file_path, 'w') as f:
                    json.dump(job_data, f, indent=2)
                print(f"Job data saved to: {file_path}")
            return job_data

        except json.JSONDecodeError:
            print(f"Failed to decode JSON for job {job_id}")
        except (socket.error, OSError) as e:
            print(f"Socket error while retrieving job data for {job_id}: {e}. Retrying...")
            time.sleep(5)
        except Exception as e:
            print(f"Error retrieving job data for {job_id}: {e}")
            break
        finally:
            if conn:
                conn.close()
            time.sleep(0.5)

    print(f"Failed to retrieve job data for job ID {job_id} after multiple attempts.")
    print(f"Job Data for {job_id}:\n{json.dumps(job_data, indent=2)}")
    return job_data


# Extract nodes (poles, anchors, etc.) from job data
def extractNodes(job_data, job_name, job_id):
    nodes = job_data.get("nodes", {})
    if not nodes:
        print("No nodes found.")
        return []

    photo_data = job_data.get('photos', {})
    trace_data_all = job_data.get('traces', {}).get('trace_data', {})
    node_points = []

    # Extract job status from job data
    job_status = job_data.get('metadata', {}).get('job_status', "Unknown")

    for node_id, node_data in nodes.items():
        attributes = node_data.get('attributes', {})
        node_type = attributes.get('node_type', {}).get('-Imported')
        has_pole_tag = 'pole_tag' in attributes
        is_pole = (node_type == 'pole' or has_pole_tag or 'pole' in node_id.lower()) and node_type != 'reference'
        is_anchor = node_type == 'new anchor'

        if is_pole or is_anchor:
            latitude = node_data.get('latitude')
            longitude = node_data.get('longitude')

            if latitude is None or longitude is None:
                print(f"Skipping node {node_id}: Missing latitude or longitude")
                continue

            # Extract additional attributes
            mr_status = "Unknown"
            if 'proposed_pole_spec' in attributes:
                mr_status = "PCO Required"
            else:
                mr_state = attributes.get('mr_state', {}).get('auto_calced', "Unknown")
                warning_present = 'warning' in attributes
                if mr_state == "No MR" and not warning_present:
                    mr_status = "No MR"
                elif mr_state == "MR Resolved" and not warning_present:
                    mr_status = "Comm MR"
                elif mr_state == "MR Resolved" and warning_present:
                    mr_status = "Electric MR"

            company = attributes.get('pole_tag', {}).get('-Imported', {}).get('company', "Unknown")
            fldcompl_value = attributes.get('field_completed', {}).get('value', "Unknown")
            fldcompl = 'yes' if fldcompl_value == 1 else 'no' if fldcompl_value == 2 else 'Unknown'
            pole_class = attributes.get('pole_class', {}).get('-Imported', "Unknown")
            pole_height = attributes.get('pole_height', {}).get('-Imported', "Unknown")
            pole_spec = attributes.get('pole_spec', {}).get('button_calced', "Unknown")
            tag = attributes.get('pole_tag', {}).get('-Imported', {}).get('tagtext', "Unknown")
            scid = attributes.get('scid', {}).get('auto_button', "Unknown")

            # Extract POA height using main photo wire data
            poa_height = ""

            # Locate the main photo
            photos = node_data.get('photos', {})
            main_photo_id = next(
                (photo_id for photo_id, photo_info in photos.items() if photo_info.get('association') == 'main'), None)

            if main_photo_id and main_photo_id in photo_data:
                photofirst_data = photo_data[main_photo_id].get('photofirst_data', {}).get('wire', {})
                for wire_id, wire_info in photofirst_data.items():
                    trace_id = wire_info.get('_trace')
                    trace_data = trace_data_all.get(trace_id, {})

                    # Check if the trace matches the desired conditions
                    if (trace_data.get('company') == 'Clearnetworx' and
                            trace_data.get('proposed', False) and
                            trace_data.get('_trace_type') == 'cable' and
                            trace_data.get('cable_type') == 'Fiber Optic Com'):

                        # Extract the measured height and convert to feet and inches
                        poa_height = wire_info.get('_measured_height')
                        if poa_height is not None:
                            feet = int(poa_height // 12)
                            inches = int(poa_height % 12)
                            poa_height = f"{feet}' {inches}\""

            # Extract POA height from "guying" if not found in wire
            if not poa_height and main_photo_id and main_photo_id in photo_data:
                guying_data = photo_data[main_photo_id].get('photofirst_data', {}).get('guying', {})
                for wire_id, wire_info in guying_data.items():
                    trace_id = wire_info.get('_trace')
                    trace_data = trace_data_all.get(trace_id, {})

                    # Check if the trace matches the desired conditions for down guy
                    if (trace_data.get('company') == 'Clearnetworx' and
                            trace_data.get('proposed', False) and
                            trace_data.get('_trace_type') == 'down_guy'):

                        # Extract the measured height and convert to feet and inches
                        poa_height = wire_info.get('_measured_height')
                        if poa_height is not None:
                            feet = int(poa_height // 12)
                            inches = int(poa_height % 12)
                            poa_height = f"{feet}' {inches}\""
                        break

            # Append the node data to the list, including job_status
            node_points.append({
                "id": node_id,
                "lat": latitude,
                "lng": longitude,
                "jobname": job_name,
                "job_status": job_status,
                "MR_statu": mr_status,
                "company": company,
                "fldcompl": fldcompl,
                "pole_class": pole_class,
                "tag": tag,
                "scid": scid,
                "POA_Height": poa_height
            })

    return node_points

def extractAnchors(job_data, job_name, job_id):
    """Extract anchor points from job data, focusing only on the Anchor_Spec attribute."""
    anchors = job_data.get("nodes", {})
    anchor_points = []

    for node_id, node_data in anchors.items():
        # Check if node_type is 'new anchor'
        if node_data.get("attributes", {}).get("node_type", {}).get("button_added") == "new anchor":
            latitude = node_data.get("latitude")
            longitude = node_data.get("longitude")

            # Extract anchor_spec
            anchor_spec_data = node_data.get("attributes", {}).get("anchor_spec", {})
            anchor_spec = anchor_spec_data.get("button_added", "Unknown")

            # Append anchor information with Anchor_Spec and Job_Name
            anchor_points.append({
                "longitude": longitude,
                "latitude": latitude,
                "anchor_spec": anchor_spec,
                "job_id": job_id
            })

    return anchor_points
# Extract connections (lines, cables, etc.) from job data

# Extract connections (lines, cables, etc.) from job data
def extractConnections(job_data, job_name, job_id):
    connections = job_data.get("connections", {})
    nodes = job_data.get("nodes", {})
    photos = job_data.get("photos", {})

    if not isinstance(connections, dict) or not nodes:
        print("Unexpected structure in 'connections' or no nodes found.")
        return []

    line_connections = []

    for conn_id, connection in connections.items():
        attributes = connection.get("attributes", {})
        connection_type = attributes.get("connection_type", {}).get("value") or attributes.get("connection_type", {}).get("button_added")

        if connection_type is None:
            print(f"Connection ID {conn_id} has no 'connection_type' attribute. Full attributes: {attributes}")
            connection_type = "Unknown"

        # Skip "reference" connection types
        if connection_type.lower() == "reference":
            continue

        #print(f"Processing Connection ID {conn_id} with connection type: {connection_type}")

        # Extract start and end node IDs
        start_node_id = connection.get("node_id_1")
        end_node_id = connection.get("node_id_2")

        if not start_node_id or not end_node_id:
            print(f"Connection ID {conn_id} is missing start or end node ID.")
            continue

        start_node = nodes.get(start_node_id)
        end_node = nodes.get(end_node_id)

        if not start_node or not end_node:
            print(f"Missing node data for connection ID {conn_id}: start_node_id={start_node_id}, end_node_id={end_node_id}")
            continue

        # Extract start and end coordinates
        start_coords = (start_node.get("longitude"), start_node.get("latitude"))
        end_coords = (end_node.get("longitude"), end_node.get("latitude"))

        if not all(start_coords) or not all(end_coords):
            print(f"Invalid coordinates for connection ID {conn_id}")
            continue

        # Initialize mid_ht to None
        mid_ht = None

        # Only apply mid_ht logic to "aerial" connection type
        if connection_type.lower() == "aerial cable":
            # Find the main photo associated with this connection in the midpoint_section
            main_photo_id = None
            sections = connection.get("sections", {}).get("midpoint_section", {})
            photos_dict = sections.get("photos", {})

            # Loop through photos and find the one with association "main"
            for photo_id, photo_details in photos_dict.items():
                if photo_details.get("association") == "main":
                    main_photo_id = photo_id
                    #print(f"Found main photo ID for connection {conn_id}: {main_photo_id}")
                    break

            if main_photo_id and main_photo_id in photos:
                main_photo_details = photos.get(main_photo_id, {})
                photofirst_entry = main_photo_details.get("photofirst_data", {})

                wires = photofirst_entry.get("wire", {})
                matching_trace_id = None

                for wire_id, wire_info in wires.items():
                    trace_id = wire_info.get("_trace")
                    trace_data = job_data.get("traces", {}).get("trace_data", {}).get(trace_id, {})
                    company = trace_data.get("company")
                    proposed = trace_data.get("proposed")

                    if company == "Clearnetworx" and proposed:
                        matching_trace_id = trace_id
                        break

                if matching_trace_id:
                    for wire_id, wire_info in wires.items():
                        if wire_info.get("_trace") == matching_trace_id:
                            mid_ht = wire_info.get("_measured_height")
                            if mid_ht is not None:
                                feet = int(mid_ht // 12)
                                inches = int(mid_ht % 12)
                                mid_ht = f"{feet}' {inches}\""
                            break

        # Append connection to list
        line_connections.append({
            "StartX": start_coords[0],
            "StartY": start_coords[1],
            "EndX": end_coords[0],
            "EndY": end_coords[1],
            "ConnType": connection_type,
            "JobName": job_name,
            "job_id": job_id,
            "mid_ht": mid_ht
        })

        #print(f"Connection ID: {conn_id}, mid_ht: {mid_ht}")

    print(f"Total connections extracted: {len(line_connections)}")
    return line_connections



def savePointsToShapefile(points, filename):
    workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [Point(point["lng"], point["lat"]) for point in points]

    gdf = gpd.GeoDataFrame(points, geometry=geometries, crs="EPSG:4326")

    # Rename columns
    gdf.rename(columns={
        'company': 'utility',
        'tag': 'pole tag',
        'fldcompl': 'collected',
        'jobname': 'jobname',
        "job_status': 'job_status',"
        'MR_statu': 'mr_status',
        'pole_spec': 'pole_spec',
        'POA_Height': 'att_ht',
        'lat': 'latitude',
        'lng': 'longitude'
    }, inplace=True)

    # Remove unwanted columns, ignore if they don't exist
    gdf.drop(columns=['pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving GeoPackage: {e}")


# Function to save line connections to a GeoPackage
def saveAnchorsToGeoPackage(anchor_points, filename):
    workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in anchor_points]

    gdf = gpd.GeoDataFrame(anchor_points, geometry=geometries, crs="EPSG:4326")

    # Rename columns
    gdf.rename(columns={
        'longitude': 'longitude',
        'latitude': 'latitude',
        'anchor_spec': 'anchor_spec'
    }, inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, layer='anchors', driver="GPKG")
        print(f"Anchors GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving anchors GeoPackage: {e}")
def saveLineShapefile(line_connections, filename):
    workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [
        LineString([(line["StartX"], line["StartY"]), (line["EndX"], line["EndY"])]
    ) for line in line_connections]

    gdf = gpd.GeoDataFrame(line_connections, geometry=geometries, crs="EPSG:4326")
    gdf.drop(columns=['StartX', 'StartY', 'EndX', 'EndY', 'job_id'], errors='ignore', inplace=True)
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"Line GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving line GeoPackage: {e}")

# Function to save nodes to a GeoPackage
def saveMasterGeoPackage(all_nodes, all_connections, all_anchors, filename):
    workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))

    # Save nodes as point layer
    if all_nodes:
        node_geometries = [Point(node["lng"], node["lat"]) for node in all_nodes]
        gdf_nodes = gpd.GeoDataFrame(all_nodes, geometry=node_geometries, crs="EPSG:4326")

        # Rename columns
        gdf_nodes.rename(columns={
            'company': 'utility',
            'tag': 'pole tag',
            'fldcompl': 'collected',
            'jobname': 'jobname',
            'job_status': 'job_status',
            'MR_statu': 'mr_status',
            'pole_spec': 'pole_spec',
            'POA_Height': 'att_ht',
            'lat': 'latitude',
            'lng': 'longitude'
        }, inplace=True)

        # Remove unwanted columns
        gdf_nodes.drop(columns=['pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)

        if not gdf_nodes.empty:
            try:
                gdf_nodes.to_file(file_path, layer='poles', driver="GPKG", mode='w', OVERWRITE="YES")
                print(f"Nodes layer successfully saved to: {file_path}")
            except Exception as e:
                print(f"Error saving nodes layer to GeoPackage: {e}")

    # Save connections as line layer
    if all_connections:
        line_geometries = [
            LineString([(line["StartX"], line["StartY"]), (line["EndX"], line["EndY"])])
            for line in all_connections
        ]
        gdf_connections = gpd.GeoDataFrame(all_connections, geometry=line_geometries, crs="EPSG:4326")

        # Drop unnecessary columns from connections
        gdf_connections.drop(columns=['StartX', 'StartY', 'EndX', 'EndY', 'job_id'], errors='ignore', inplace=True)

        if not gdf_connections.empty:
            try:
                gdf_connections.to_file(file_path, layer='connections', driver="GPKG", mode='w', OVERWRITE="YES")
                print(f"Connections layer successfully saved to: {file_path}")
            except Exception as e:
                print(f"Error saving connections layer to GeoPackage: {e}")

    # Save anchors as point layer
    if all_anchors:
        anchor_geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in all_anchors]
        gdf_anchors = gpd.GeoDataFrame(all_anchors, geometry=anchor_geometries, crs="EPSG:4326")

        # Rename columns
        gdf_anchors.rename(columns={
            'longitude': 'longitude',
            'latitude': 'latitude',
            'anchor_spec': 'anchor_spec'
        }, inplace=True)

        if not gdf_anchors.empty:
            try:
                gdf_anchors.to_file(file_path, layer='anchors', driver="GPKG", mode='w', OVERWRITE="YES")
                print(f"Anchors layer successfully saved to: {file_path}")
            except Exception as e:
                print(f"Error saving anchors layer to GeoPackage: {e}")
        # Overwrite the Feature Service in ArcGIS Enterprise
    overwriteFeatureService(file_path)

# Function to save line connections to a GeoPackage
def saveMasterConnectionsToGeoPackage(all_connections, filename):
    workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))

    # Create geometries for each connection using StartX, StartY, EndX, and EndY
    geometries = [
        LineString([(line["StartX"], line["StartY"]), (line["EndX"], line["EndY"])])
        for line in all_connections
    ]

    # Create GeoDataFrame including mid_ht field
    gdf = gpd.GeoDataFrame(all_connections, geometry=geometries, crs="EPSG:4326")

    # Ensure mid_ht field is explicitly present in the DataFrame
    if 'mid_ht' not in gdf.columns:
        gdf['mid_ht'] = None  # Add the column if it doesn't exist

    # Drop unnecessary columns from connections
    gdf.drop(columns=['StartX', 'StartY', 'EndX', 'EndY', 'job_id'], errors='ignore', inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"Master connections GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving master connections GeoPackage: {e}")

def saveMasterGeoPackage(all_nodes, all_connections, all_anchors, filename):
    workspace_path = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))

    # Save nodes as point layer
    if all_nodes:
        date_stamp = datetime.now().strftime("%Y-%m-%d")
        node_geometries = [Point(node["lng"], node["lat"]) for node in all_nodes]
        filename = f"Master_{date_stamp}.gpkg"
        gdf_nodes = gpd.GeoDataFrame(all_nodes, geometry=node_geometries, crs="EPSG:4326")

        # Rename columns
        gdf_nodes.rename(columns={
            'company': 'utility',
            'tag': 'pole tag',
            'fldcompl': 'collected',
            'jobname': 'jobname',
            'job_status': 'job_status',
            'MR_statu': 'mr_status',
            'pole_spec': 'pole_spec',
            'POA_Height': 'att_ht',
            'lat': 'latitude',
            'lng': 'longitude'
        }, inplace=True)

        # Remove unwanted columns
        gdf_nodes.drop(columns=['pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)

        if not gdf_nodes.empty:
            try:
                gdf_nodes.to_file(file_path, layer='poles', driver="GPKG", mode='w', OVERWRITE="YES")
                print(f"Nodes layer successfully saved to: {file_path}")
            except Exception as e:
                print(f"Error saving nodes layer to GeoPackage: {e}")

    # Save connections as line layer
    if all_connections:
        line_geometries = [
            LineString([(line["StartX"], line["StartY"]), (line["EndX"], line["EndY"])])
            for line in all_connections
        ]
        gdf_connections = gpd.GeoDataFrame(all_connections, geometry=line_geometries, crs="EPSG:4326")

        # Drop unnecessary columns from connections
        gdf_connections.drop(columns=['StartX', 'StartY', 'EndX', 'EndY', 'job_id'], errors='ignore', inplace=True)

        if not gdf_connections.empty:
            try:
                gdf_connections.to_file(file_path, layer='connections', driver="GPKG", mode='w', OVERWRITE="YES")
                print(f"Connections layer successfully saved to: {file_path}")
            except Exception as e:
                print(f"Error saving connections layer to GeoPackage: {e}")

    # Save anchors as point layer
    if all_anchors:
        anchor_geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in all_anchors]
        gdf_anchors = gpd.GeoDataFrame(all_anchors, geometry=anchor_geometries, crs="EPSG:4326")

        # Rename columns
        gdf_anchors.rename(columns={
            'longitude': 'longitude',
            'latitude': 'latitude',
            'anchor_spec': 'anchor_spec'
        }, inplace=True)

        if not gdf_anchors.empty:
            try:
                gdf_anchors.to_file(file_path, layer='anchors', driver="GPKG", mode='w', OVERWRITE="YES")
                print(f"Anchors layer successfully saved to: {file_path}")
            except Exception as e:
                print(f"Error saving anchors layer to GeoPackage: {e}")


# Function to create a report of MR Status counts per job
def create_report(jobs_summary):
    report_data = []

    for job in jobs_summary:
        job_name = job['job_name']

        # Extract and clean job status
        job_status = job.get('job_status', 'Unknown').strip()

        mr_status_counts = job['mr_status_counts']
        pole_count = sum(mr_status_counts.values())

        report_data.append({
            'Job Name': job_name,
            'Job Status': job_status,
            'No MR': mr_status_counts.get('No MR', 0),
            'Comm MR': mr_status_counts.get('Comm MR', 0),
            'Electric MR': mr_status_counts.get('Electric MR', 0),
            'PCO Required': mr_status_counts.get('PCO Required', 0),
            'Pole Count': pole_count
        })

    # Create a DataFrame from the report data
    df_report = pd.DataFrame(report_data)

    # Ensure the directory exists
    workspace_dir = r"C:\Users\lewis\Documents\Deeply_Digital\Katapult_Automation\workspace"
    if not os.path.exists(workspace_dir):
        try:
            os.makedirs(workspace_dir)
            print(f"Workspace directory created: {workspace_dir}")
        except Exception as e:
            print(f"Failed to create workspace directory: {e}")
            return None

    # Generate a filename with a timestamp
    timestamp = datetime.now().strftime("%m%d%Y_%I%M")
    report_filename = f"Aerial_Status_Report_{timestamp}.xlsx"
    report_path = os.path.join(workspace_dir, report_filename)

    # Write the report to an Excel file with formatting
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Aerial Status Report"

        # Add merged header with title in the first row
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Aerial Status Report"
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add the date/time in the second row
        ws.merge_cells('A2:G2')
        date_cell = ws.cell(row=2, column=1)
        date_cell.value = datetime.now().strftime('%m/%d/%Y %I:%M:%p')
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set row height for the first row to accommodate title
        ws.row_dimensions[1].height = 30
        # Set row height for the second row to accommodate the date
        ws.row_dimensions[2].height = 20

        # Add the column headers with styling in the third row
        for col_num, column_title in enumerate(df_report.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Color the column headers according to their field
            header_colors = {
                "Job Name": "CCFFCC",
                "Job Status": "CCFFCC",
                "No MR": "D9D9D9",
                "Comm MR": "FFFF00",
                "Electric MR": "FFC000",
                "PCO Required": "FF0000",
                "Pole Count": "CCFFCC",
            }
            if column_title in header_colors:
                cell.fill = PatternFill(start_color=header_colors[column_title], end_color=header_colors[column_title],
                                        fill_type="solid")

        # Add the data rows
        for r_idx, row in enumerate(dataframe_to_rows(df_report, index=False, header=False), 4):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Adjust column widths: "Job Status" column to 23.71 and others to 13.3
        for col in ws.iter_cols(min_row=3, max_row=3):  # Iterating over the column headers row
            column_letter = col[0].column_letter
            if col[0].value == "Job Status":
                ws.column_dimensions[column_letter].width = 23.71
            elif col[0].value == "Job Name":
                ws.column_dimensions[column_letter].width = 44
            else:
                ws.column_dimensions[column_letter].width = 13.3

        # Add borders around the entire report including title and date rows
        all_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Apply borders to all cells including title and date rows
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = all_border

        # Add Job Status Summary Headers in Two Rows (4 statuses per row) with colors
        job_statuses_row_1 = [
            ("Pending Field Collection", "CCFFCC"),
            ("Pending Photo Annotation", "B7DEE8"),
            ("Sent to PE", "CCC0DA"),
            ("Pending EMR", "FFC000")
        ]
        job_statuses_row_2 = [
            ("Approved for Construction", "9BBB59"),
            ("Hold", "BFBFBF"),
            ("As Built", "FABF8F"),
            ("Delivered", "92D050")
        ]

        # Add first row of job statuses from I6 to L6
        for col_num, (status, color) in enumerate(job_statuses_row_1, 9):  # Start at column 'I' (index 9)
            header_cell = ws.cell(row=6, column=col_num)
            header_cell.value = status
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            # Set the column width to 24.14
            ws.column_dimensions[header_cell.column_letter].width = 24.14

        job_status_counts = {status[0]: 0 for status in job_statuses_row_1 + job_statuses_row_2}

        for job in jobs_summary:
            job_status = job.get('job_status', 'Unknown').strip()  # Consistent retrieval as in extractNodes
            if job_status in job_status_counts:
                job_status_counts[job_status] += 1

        # Write the counts in row 7 under each corresponding header for the first row of statuses
        for col_num, (status, _) in enumerate(job_statuses_row_1, 9):  # Start at column 'I' (index 9)
            count_cell = ws.cell(row=7, column=col_num)
            count_cell.value = job_status_counts[status]
            count_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add second row of job statuses from I9 to L9
        for col_num, (status, color) in enumerate(job_statuses_row_2, 9):  # Start at column 'I' (index 9)
            header_cell = ws.cell(row=9, column=col_num)
            header_cell.value = status
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write the counts in row 10 under each corresponding header for the second row of statuses
        for col_num, (status, _) in enumerate(job_statuses_row_2, 9):  # Start at column 'I' (index 9)
            count_cell = ws.cell(row=10, column=col_num)
            count_cell.value = job_status_counts[status]
            count_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply borders to the new summary table headers and counts
        for row in ws.iter_rows(min_row=6, max_row=7, min_col=9,
                                max_col=12):  # Columns I to L (indices 9 to 12) for the first row
            for cell in row:
                cell.border = all_border
        for row in ws.iter_rows(min_row=9, max_row=10, min_col=9,
                                max_col=12):  # Columns I to L (indices 9 to 12) for the second row
            for cell in row:
                cell.border = all_border

        # Save the workbook
        wb.save(report_path)
        print(f"Report successfully created: {report_path}")
    except Exception as e:
        print(f"Error creating report: {e}")

    return report_path

# Function to send email notification with attachment
def send_email_notification(email_list, report_path):
    try:
        smtp_server = "smtp.office365.com"  # Your SMTP server
        smtp_port = 587
        smtp_user = "brandan.lewis@deeplydigital.com"  # Your email address
        smtp_password = "Bmxican123!"  # Your email password

        # Set up the SMTP server
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)

        # Create the email
        from_email = smtp_user
        to_emails = email_list

        for to_email in to_emails:
            msg = MIMEMultipart('alternative')
            msg['From'] = from_email
            msg['To'] = to_email
            msg['Subject'] = f"Aerial Status Report: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"

            # Plain text version of the email body
            text_body = (
                "Hey Team,\n\n"
                "The Katapult API automation script has finished running and the report has been generated.\n"
                "The report is designed to give an overview of all poles that are in design with metrics to support decision making and cost planning.\n"
                "Please find the attached report for more details.\n\n"
                "Thanks,\n"
                "Brandan"
            )

            # HTML version of the email body
            html_body = """
            <html>
            <body>
                <p>Hey Team,</p>
                <p>The Katapult automation script has <b>finished running</b> and the report has been generated.</p>
                <p>The report is designed to give an overview of all poles that are in design with metrics to support decision making and cost planning.</p>
                <p>See the updated layer in the web map below.</p>
                <p><a href='https://gis.clearnetworx.com/portal/apps/mapviewer/index.html?webmap=e3ee3bfdc6184987baf87f9f0ebd23ef'>Katapult Master API Map</a></p>
                <p>Please find the attached report for more details.</p>
                <p>Thanks,<br>
                   <i>Brandan</i>
                </p>
            </body>
            </html>
            """

            # Attach both plain text and HTML versions to support different email clients
            part1 = MIMEText(text_body, 'plain')
            part2 = MIMEText(html_body, 'html')

            msg.attach(part1)
            msg.attach(part2)

            # Retry mechanism to wait for the report file to be available
            max_retries = 5
            retries = 0
            while not os.path.exists(report_path) and retries < max_retries:
                print(f"Waiting for report file to be available: {report_path} (Attempt {retries + 1})")
                time.sleep(2)  # Wait for 2 seconds before retrying
                retries += 1

            # Attach the Excel file if it exists
            if os.path.exists(report_path):
                attachment = MIMEBase('application', 'octet-stream')
                try:
                    with open(report_path, "rb") as attachment_file:
                        attachment.set_payload(attachment_file.read())
                    encoders.encode_base64(attachment)
                    attachment.add_header('Content-Disposition', f'attachment; filename={os.path.basename(report_path)}')
                    msg.attach(attachment)
                except Exception as e:
                    print(f"Error reading the report file for attachment: {e}")
                    continue
            else:
                print(f"Error: Report file not found after {max_retries} retries. Skipping email attachment.")
                continue

            # Send the email
            server.send_message(msg)
            print(f"Email sent to {to_email}")

        # Close the SMTP server
        server.quit()

    except Exception as e:
        print(f"Error sending email: {e}")


# Main function to run the job for testing
def main(email_list):

    all_jobs = []

    if TEST_ONLY_SPECIFIC_JOB:
        all_jobs = [{'id': TEST_JOB_ID, 'name': 'Test Job'}]
    else:
        all_jobs = getJobList()

    all_nodes = []
    all_connections = []
    all_anchors = []
    jobs_summary = []

    if not all_jobs:
        print("No jobs found.")
        return

    for job in all_jobs:
        job_id = job['id']
        job_name = job['name']
        print(f"Processing job: {job_name} (ID: {job_id})")

        job_data = getJobData(job_id)

        if job_data:
            nodes = extractNodes(job_data, job_name, job_id)
            connections = extractConnections(job_data, job_name, job_id)
            anchors = extractAnchors(job_data, job_name, job_id)

            if nodes:
                all_nodes.extend(nodes)

                # Summarize MR Status counts for the job
                mr_status_counts = {}
                for node in nodes:
                    mr_status = node['MR_statu']
                    if mr_status not in mr_status_counts:
                        mr_status_counts[mr_status] = 0
                    mr_status_counts[mr_status] += 1

                jobs_summary.append({
                    'job_name': job_name,
                    'job_status': job_data.get('metadata', {}).get('job_status', 'Unknown'),  # Extract job status from job_data,
                    'mr_status_counts': mr_status_counts
                })

            if connections:
                all_connections.extend(connections)
            if anchors:
                all_anchors.extend(anchors)

    # Only save if data is present
    if all_nodes or all_connections or all_anchors:
        # Save all nodes, connections, and anchors to master GeoPackages
        saveMasterGeoPackage(all_nodes, all_connections, all_anchors, "Master.gpkg")
    else:
        print("No data extracted for any job. Nothing to save.")

    # Create a report with MR Status counts per job
    report_path = None
    if jobs_summary:
        report_path = create_report(jobs_summary)

    # Send email notification after report generation
    if report_path:
        send_email_notification(email_list, report_path)

if __name__ == "__main__":
    # Email list to notify when the report is done

    email_list = ["brandan.lewis@deeplydigital.com"]
    start_time = time.time()  # Record the start time
    main(email_list)
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Total execution time: {elapsed_time:.2f} seconds")
